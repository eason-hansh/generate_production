#!/usr/bin/env python3
"""
PDF信息提取模块
使用marker和dashscope(Qwen API)提取PDF中的采购订单信息
"""

import json
import re
from pathlib import Path
from typing import Dict, Any, Optional, List, Union

import dashscope
from http import HTTPStatus
from openpyxl import load_workbook

from marker.converters.pdf import PdfConverter
from marker.models import create_model_dict
from marker.output import text_from_rendered


dashscope.api_key = 'sk-349857a349fe47adb358f784c7860d6a'

class PDFExtractor:
    """PDF信息提取器"""
    
    def __init__(self):
        """初始化PDF提取器"""

        self.converter = PdfConverter(
            artifact_dict=create_model_dict(),
        )
        # logger.info("pdf converter started")
        # 设计提示词提取 PO 中的 po_no 和 采购商品的货号、数量
        ai_output_template = {
            "po_no": "提取的采购订单号",
            "product_info": [
                {"cust_item_code": "08484", "quantity": "50000"},
                {"cust_item_code": "08485", "quantity": "6000"}
            ]
        }
        self.prompt = (f"你是一位专业的信息采集员，擅长从文本中精准提取指定字段并以JSON格式返回。请严格遵循以下要求：\n"
                       f"## 任务\n"
                       f"从输入的采购订单中提取：\n"
                       f"1. 采购订单号（po_no）：识别 'PURCHASE ORDER NUMBER'、'PO'等关键字后的字母数字组合；\n"
                       f"2. 所有采购商品信息：每个商品的'货号(custom item code)'和'数量(quantity)'。\n"
                       f"## 输出格式\n"
                       f"{ai_output_template}\n"
                       f"## 强制规则\n"
                       f"1. 只返回JSON格式的数据，不要包含任何额外的解释或文本；\n"
                       f"2. 若无法找到po_no，该字段值为空字符串；\n"
                       f"3. '货号(custom item code)' 仅考虑 'EPM Part'对应的信息；\n"
                       f"4. product_info必须是数组，即使只有1个商品；\n"
                       f"5. quantity字段必须为整数类型（无逗号、无小数点）；\n")

    def parse_pdf(self, pdf_path):
        rendered = self.converter(pdf_path)
        text, _, images = text_from_rendered(rendered)

        return text

    @staticmethod
    def call_api(prompt):
        messages = [
            {'role': 'user', 'content': prompt}
        ]
        response = dashscope.Generation.call(
            'qwen-plus',
            messages=messages,
            temperature=0.1,
            result_format='message'  # set the result is message format.
        )
        # responses = dashscope.Generation.call(
        #     model="qwen-plus",
        #     api_key="sk-349857a349fe47adb358f784c7860d6a",
        #     messages=messages,
        #     stream=True,
        #     result_format='message',  # 将返回结果格式设置为 message
        #     top_p=0.8,
        #     temperature=0.1,
        #     enable_search=False,
        #     enable_thinking=False,
        #     thinking_budget=1000,
        # )
        # for response in responses:
        #     a = 1
        if response.status_code == HTTPStatus.OK:
            output = response['output']['choices'][0]['message']['content']
            input_tokens = response['usage']['input_tokens']
            output_tokens = response['usage']['output_tokens']
            return output, input_tokens, output_tokens

        else:
            print('Request id: %s, Status code: %s, error code: %s, error message: %s' % (
                response.request_id, response.status_code,
                response.code, response.message
            ))
            return "{}", 0, 0

    def extract_customer_codes_from_excel(self, excel_path: str) -> List[str]:
        """
        从Excel 主表 中提取客户货号列表
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            客户货号列表
            
        Raises:
            Exception: 当文件读取失败或工作表不存在时抛出异常
        """
        try:
            # 加载Excel文件
            wb = load_workbook(filename=excel_path, data_only=True)
            
            # 检查工作表是否存在，并处理可能的名称问题
            sheet_names = wb.sheetnames
            main_sheet_name = None
            
            # 查找主表（可能包含空格或特殊字符）
            for sheet_name in sheet_names:
                if "主表" in sheet_name or sheet_name.strip() == "主表":
                    main_sheet_name = sheet_name
                    break
            
            if not main_sheet_name:
                raise ValueError(f"未找到主表工作表。可用工作表: {sheet_names}")
            
            main_sheet = wb[main_sheet_name]
            
            customer_codes = []
            exit_flag = False
            
            # 遍历行，找到"客户货号"标记后的所有货号
            for row in main_sheet.iter_rows():
                try:
                    if exit_flag and row[0].value:
                        # 提取货号（A列的值）
                        customer_code = str(row[0].value).strip()
                        if customer_code and customer_code != "客户货号":
                            customer_codes.append(customer_code)
                    
                    # 找到"客户货号"标记
                    if row[0].value and "客户货号" in str(row[0].value):
                        exit_flag = True
                        continue
                        
                except Exception as e:
                    print(f"处理Excel行时出错: {e}")
                    continue
            
            return customer_codes
            
        except Exception as e:
            raise Exception(f"从Excel文件 {excel_path} 提取客户货号失败: {e}")

    def extract_info(self, parsed_pdf, customer_codes_reference: str = ""):
        # 基于 qwen api,以提示词的方式抽取信息
        prompt = (f"{self.prompt}{customer_codes_reference}\n"
                  f"## 待处理采购订单：\n{parsed_pdf}")
        raw_production_info, input_tokens, output_tokens = self.call_api(prompt)
        # logger.info(f"raw_ai_result: {raw_production_info}")
        # logger.info(f"input_token_nums: {input_tokens}")
        # logger.info(f"output_token_nums: {output_tokens}")

        return raw_production_info

    @staticmethod
    def convert_json_2_dict(raw_json):
        # 提取 {}里的所有内容，再转为字典
        start_index = raw_json.find('{')
        if start_index == -1:
            raise ValueError("未检测到左大括号 {")

        # 定位最后一个右大括号
        end_index = raw_json.rfind('}')
        if end_index == -1:
            raise ValueError("未检测到右大括号 }")

        # 检查括号顺序有效性
        if end_index <= start_index:
            raise ValueError("右大括号位置在左大括号之前或相同")

        # 提取目标子串
        json_str = raw_json[start_index:end_index + 1]
        # 将所有的单引号 --> 双引号
        new_json_str = json_str.replace("'", '"')

        try:
            # 尝试解析JSON
            json_2_dict = json.loads(new_json_str)
            # logger.info(f"ai_output: {json_2_dict}")
            return json_2_dict

        except json.JSONDecodeError as e:
            # 增强错误信息提示
            error_msg = f"JSON解析失败: {e.msg}\n错误位置: {e.pos}\n错误片段: {new_json_str[max(0, e.pos - 20):e.pos + 20]}"
            raise ValueError(error_msg) from e

    def process(self, pdf_path, templates: Dict[str, str]):
        """
        pdf处理流程如下：
        1. 解析 pdf --> markdown
        2. ai extract raw_info
        3. raw info --> dict
        """
        pdf_2_markdown = self.parse_pdf(pdf_path)

        # 抽取模板中的客户货号，作为 AI 的 reference;
        # template_2_info 中包含模板的：位置+客户货号
        customer_codes_reference, template_2_info = self.get_customer_codes_reference(templates)
        # ai 抽取信息
        product_info_str = self.extract_info(pdf_2_markdown, customer_codes_reference)
        # 解析 ai 的生成结果
        po_product_info_dict = self.convert_json_2_dict(product_info_str)

        # po_product_info_dict = {'po_no': '4647647606', 'product_info': [{'cust_item_code': '22YM05', 'quantity': 200}, {'cust_item_code': '22YM06', 'quantity': 200}, {'cust_item_code': '23PC31', 'quantity': 200}, {'cust_item_code': '23PC32', 'quantity': 280}, {'cust_item_code': '23PC34', 'quantity': 120}, {'cust_item_code': '23PC35', 'quantity': 200}, {'cust_item_code': '23PC37', 'quantity': 240}, {'cust_item_code': '23PC38', 'quantity': 300}, {'cust_item_code': '23PC39A', 'quantity': 300}, {'cust_item_code': '23PC40A', 'quantity': 1000}, {'cust_item_code': '23PC41', 'quantity': 200}, {'cust_item_code': '23PC43', 'quantity': 1000}, {'cust_item_code': '23PC96', 'quantity': 100}, {'cust_item_code': '450G83A', 'quantity': 100}, {'cust_item_code': '5RXF2A', 'quantity': 200}, {'cust_item_code': '5RXF3B', 'quantity': 100}, {'cust_item_code': '5RXF6A', 'quantity': 200}, {'cust_item_code': '5RXF7B', 'quantity': 100}, {'cust_item_code': '5RXF8B', 'quantity': 100}, {'cust_item_code': '5RXF9B', 'quantity': 100}, {'cust_item_code': '5RXG1A', 'quantity': 200}, {'cust_item_code': '5RXG2B', 'quantity': 100}, {'cust_item_code': '5RXG3A', 'quantity': 500}, {'cust_item_code': '5RXG4A', 'quantity': 112}, {'cust_item_code': '5RXG5A', 'quantity': 500}, {'cust_item_code': '5RXG6A', 'quantity': 112}, {'cust_item_code': '5RXG7A', 'quantity': 112}]}
        # template_2_info = {'GM': {'customer_codes': ['450G83A', '22YM06', '22YM05', '22YM07', '34NK65'], 'template_path': 'company_templates/522/522_广美_任务单.xlsx'}, 'GX': {'customer_codes': ['23PC31', '23PC32', '23PC33', '23PC34', '23PC35', '23PC36', '23PC37', '23PC38', '23PC39A', '23PC40A', '23PC41', '23PC42', '23PC43', '23PC44', '23PC45', '23PC46', '23PC95', '23PC96', '23PC97', '23PC98', '5RXF2A', '5RXF3B', '5RXF4B', '5RXF5A', '5RXF6A', '5RXF7B', '5RXF8B', '5RXF9B', '5RXG1A', '5RXG2B', '5RXG3A', '5RXG4A', '5RXG5A', '5RXG6A', '5RXG7A', '5RXG0'], 'template_path': 'company_templates/522/522_广线_任务单.xlsx'}}

        # 检测 product_info 是否为空或无效
        if not po_product_info_dict.get('product_info') or len(po_product_info_dict.get('product_info', [])) == 0:
            raise ValueError("AI 无法提取 PO 中的产品信息，请手动制作生产任务单。")

        return po_product_info_dict, template_2_info

    def get_customer_codes_reference(self, templates: Dict[str, str]):
        """
        获取excel中，客户货号，作为参考字符串，辅助AI进行抽取
        
        Args:
            templates: 模板字典 {'GX': 'path1', 'GM': 'path2'}
            
        Returns:
            1. 格式化的参考货号字符串
            2. template_2_info， 包含：模板的位置信息 + 客户货号
        """
        all_customer_codes = []
        template_2_info = {}
        # 遍历所有模板路径
        for template_type, template_path in templates.items():
            # 从模板中得到 客户货号
            customer_codes = self.extract_customer_codes_from_excel(template_path)
            all_customer_codes.extend(customer_codes)
            template_2_info[template_type] = {
                'template_path': template_path,
                'customer_codes': customer_codes,
            }
        
        return self.format_reference_codes(all_customer_codes), template_2_info

    def format_reference_codes(self, codes: List[str]) -> str:
        """
        格式化参考货号字符串
        
        Args:
            codes: 客户货号列表
            
        Returns:
            格式化的参考货号字符串
        """
        if not codes:
            return ""
        
        codes_str = ", ".join([f'"{code}"' for code in codes])
        return f"\n## 参考货号格式\n根据Excel模板中的客户货号，货号格式参考如下：\n[{codes_str}]\n\n请确保提取的货号格式与上述参考格式保持一致。\n"