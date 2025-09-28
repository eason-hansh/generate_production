import json
import asyncio
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from torch.autograd.graph import saved_tensors_hooks


class ExcelProcessor:
    """
    Excel处理类，负责生成生产任务单

    ai_output_template = {
            "po_no": "提取的采购订单号",
            "product_info": [
                {"cust_item_code": "08484", "quantity": "50000"},
                {"cust_item_code": "08485", "quantity": "6000"}
            ]
        }
    """
    
    def __init__(self):
        pass

    def extract_tl_items_from_tl_sheet(self, wb):
        """
        从TL工作表中提取TL客户货号

        Args:
            wb: Excel工作簿对象

        Returns:
            TL客户货号集合
        """
        try:
            if 'TL' not in wb.sheetnames:
                return []

            tl_sheet = wb['TL']
            tl_items = []
            found_data_start = False

            # 遍历TL工作表，找到"客户货号"开始，"封箱/打包方式"结束
            for row in tl_sheet.iter_rows():
                cell_value = str(row[0].value).strip() if row[0].value else ""

                if cell_value:  # A列有值
                    if not found_data_start:  # 还没开始提取数据
                        if "客户货号" in cell_value:  # 这是开始信号
                            found_data_start = True
                            continue
                    else:  # 已经开始提取数据
                        if "封箱/打包方式" in cell_value:  # 这是结束信号
                            break
                        else:  # 这是数据行
                            tl_items.append(cell_value)
                elif found_data_start:  # 已经开始提取数据，但A列为空，停止提取
                    break

            return tl_items

        except Exception as e:
            raise Exception(f"从TL工作表提取客户货号失败: {e}")

    def generate_task_order_no(self, raw_pdf_info, task_order_no, tl_pdf_items, bc_pdf_items):
        # 批量生成 任务单
        left = task_order_no.find("(1)")
        if left == -1:
            raise ValueError("输入字符串必须包含 '(1)'")

        # 拆分前缀和后缀
        prefix = task_order_no[:left]  # "TW25040782"
        suffix = task_order_no[left + 3:]  # "BC"（跳过 "(1)"）

        product_num = len(raw_pdf_info['product_info']) # 本次PO单中的采购商品数
        # 构造一个仅包含：本次采购商品货号 + 数量 + 生产任务单号 的信息，便于直接填入 excel 模板
        pdf_info = {}
        if tl_pdf_items:  # 存在 tl sheet，则 suffix 是 TL/BC
            counter = 1
            for index, item in enumerate(raw_pdf_info['product_info']):
                if item['cust_item_code'] in tl_pdf_items:
                    pdf_info[item['cust_item_code']] = {
                        'quantity': item['quantity'],
                        'task_order_no': f"{prefix}({counter})TL"
                    }
                elif item['cust_item_code'] in bc_pdf_items:
                    pdf_info[item['cust_item_code']] = {
                        'quantity': item['quantity'],
                        'task_order_no': f"{prefix}({counter})BC"
                    }
                else:
                    pass
                counter += 1
        else:
            # 不存在 TL，全部以 BC 结尾
            task_orders = [f"{prefix}({i}){suffix}" for i in range(1, product_num + 1)]
            for index, item in enumerate(raw_pdf_info['product_info']):
                pdf_info[item['cust_item_code']] = {
                    'quantity': item['quantity'],
                    'task_order_no': task_orders[index]
                }

        # 生成范围格式的任务单号
        if product_num == 1:
            task_order_range = f"{prefix}(1){suffix}"
        else:
            task_order_range = f"{prefix}(1-{product_num}){suffix}"

        return pdf_info, task_order_range

    def process_GM_template(self, wb, raw_pdf_info, order_date, delivery_date, task_order_no, customer_code, output_dir, customer_codes, task_id, tl_pdf_items=None):
        # 将sheet的修改部分，高亮出来，创建高亮样式（黄色背景）
        formula_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 获取主表
        main_sheet = wb['主表']
        modify_sheets = ['主表']

        # 填写基本信息填写
        main_sheet['E1'] = order_date  # 制单日期
        main_sheet['B3'] = raw_pdf_info['po_no']  # PO NO
        main_sheet['B2'] = delivery_date  # 交货期

        # 生成批量的任务单号。
        # 请注意：
        # 广美中一定不存在 TL 的情况。此时，tl_pdf_items 为 None
        # 广线中可能存在TL。如果有，TL 在 BC 前
        pdf_info, task_order_range = self.generate_task_order_no(raw_pdf_info, task_order_no, bc_pdf_items=customer_codes, tl_pdf_items=tl_pdf_items)

        # 请注意：遍历 row，到 总计的 row break
        exit_flag = False
        for row in main_sheet.iter_rows():
            try:
                if exit_flag and row[0].value:
                    product_name = str(row[0].value) if type(row[0].value) is int else row[0].value.split()[0]
                    # 请注意：遍历 row，到 "总计", break
                    if product_name == "合计":
                        break
                    row_idx = row[0].row  # 获取当前行号
                    # 匹配到产品则进行修改，否则，则将改行进行隐藏
                    if product_name in pdf_info:
                        # 修改数量
                        main_sheet[f"D{row_idx}"] = int(pdf_info[product_name]['quantity'])
                        # 修改 任务单号
                        main_sheet[f"G{row_idx}"] = pdf_info[product_name]['task_order_no']
                        main_sheet.row_dimensions[row_idx].hidden = False  # 该行一定不隐藏
                        modify_sheets.append(str(row[0].value))
                    else:
                        # 修改数量
                        main_sheet[f"D{row_idx}"] = 0
                        # 修改生产任务单
                        main_sheet[f"G{row_idx}"] = ''
                        # # 外箱流水号转为空
                        # main_sheet[f"G{row_idx}"] = ''
                        # 改行 一定 隐藏
                        main_sheet.row_dimensions[row_idx].hidden = True

                if row[0].value and "客户货号" in str(row[0].value):
                    exit_flag = True
                    continue
            except Exception as e:
                print(e)

        # 将修改后的sheet显示，其余的隐藏
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet.sheet_state = 'visible' if sheet_name in modify_sheets else 'hidden'

            # 高亮公式单元格，只处理可见工作表
            if sheet.sheet_state == 'visible':
                for row in sheet.iter_rows():
                    for cell in row:
                        # 检查公式单元格（兼容不同格式）
                        if cell.data_type == 'f' or (cell.value and str(cell.value).startswith('=')):
                            cell.fill = formula_fill

        # 如果存在 TL sheet，还需要对 TL sheet 进行修改
        if tl_pdf_items:
            # 本次从 PO 单中抽取的、属于 TL sheet 的单号
            tl_customer_codes_from_po = [i['cust_item_code'] for i in raw_pdf_info['product_info'] if i['cust_item_code'] in tl_pdf_items]
            tl_sheet = wb["TL"]

            found_data_start = False

            for row in tl_sheet.iter_rows():
                cell_value = str(row[0].value).strip() if row[0].value else ""
                row_idx = row[0].row  # 获取当前行号
                if cell_value:  # A列有值
                    if not found_data_start:
                        if "客户货号" in cell_value:  # 开始信号
                            found_data_start = True
                            continue
                    else:
                        if "封箱/打包方式" in cell_value:  # 结束信号
                            break
                        else:  # 数据行
                            if cell_value in tl_customer_codes_from_po:
                                tl_sheet.row_dimensions[row_idx].hidden = False  # 该行一定不隐藏
                            else:
                                tl_sheet.row_dimensions[row_idx].hidden = True  # 隐藏
                elif found_data_start:
                    break

            # 确保 sheet 'TL' 可见
            tl_sheet.sheet_state = "visible"

        # 生成保存路径（按客户号+时间戳分文件夹）
        if output_dir is None:
            output_dir = Path('.')
        else:
            output_dir = Path(output_dir)
        
        # 使用任务ID创建文件夹
        customer_output_dir = output_dir / f"{customer_code}_{task_id}"
        customer_output_dir.mkdir(parents=True, exist_ok=True)

        # 生成文件名：任务单号_PO号_客户号
        po_no = raw_pdf_info.get('po_no', '')
        if po_no:
            po_prefix = f"PO{po_no}"
        else:
            po_prefix = "PO"

        if customer_code:
            filename = f"{task_order_range}_{po_prefix}_{customer_code}.xlsx"
        else:
            # 如果客户号提取失败，使用原来的命名方式
            filename = f"{task_order_range}_{po_prefix}.xlsx"

        output_path = customer_output_dir / filename

        wb.save(output_path)

        return str(output_path)

    def process_GX_template(self, raw_pdf_info, template_2_info, task_order_no, order_date, delivery_date, output_dir, pdf_name, customer_code, task_id):
        """
        区分是有 TL sheet，如果没有，就是 process_GM_template
        如果有，就在此基础上，再去处理 TL sheet
        Args:
            raw_pdf_info: PDF提取的信息
            template_info: 字典形式，包含广线模板路径+客户货号 {'template_path': '', 'customer_codes': ''}
            task_order_no: 任务单号
            order_date: 制单日期
            delivery_date: 交货期
            output_dir: 输出目录
            pdf_name: PDF文件名
            customer_code: 客户号
            
        Returns:
            生成的Excel文件路径
        """
        try:
            # 1. 加载Excel模板，注意使用 data_only=False 以保留公式
            excel_path = template_2_info['GX']['template_path']
            wb = load_workbook(filename=excel_path, data_only=False)

            # 2. 检查是否存在TL工作表
            has_tl_sheet = 'TL' in wb.sheetnames

            if not has_tl_sheet:
                output_path = self.process_GM_template(wb, raw_pdf_info, order_date, delivery_date, task_order_no, customer_code, output_dir, template_2_info['GX']['customer_codes'], task_id)

            else:
                # 区分 TL 和 BC
                tl_items = self.extract_tl_items_from_tl_sheet(wb)

                all_customer_codes_from_gx = template_2_info['GX']['customer_codes']
                if not all_customer_codes_from_gx:
                    raise ValueError("未获得广线 TL sheet 中的客户货号，请检查！")

                bc_items = [c for c in all_customer_codes_from_gx if c not in tl_items] if tl_items else all_customer_codes_from_gx

                output_path = self.process_GM_template(wb, raw_pdf_info, order_date, delivery_date, task_order_no, customer_code, output_dir, bc_items, task_id, tl_pdf_items=tl_items)

            return output_path
            
        except Exception as e:
            raise Exception(f"处理广线模板失败: {e}")

    def process(self, raw_pdf_info, template_2_info, task_order_no, order_date, delivery_date, output_dir,
                pdf_name, customer_code, task_id):
        """
        将抽取后的信息填写入 Excel模板（支持单模板和双模板）

        GM 的基础模板可以沿用上一版本的填充模板
        GX 的分为是否有 TL sheet，如果没，和GM一致；如果有，先使用 GM 处理方式，再对 TL sheet 进行额外处理

        Args:
            raw_pdf_info: PDF提取的信息
            template_2_info: 模板字典 {'GX': {'template_path': '', 'customer_codes': ''}}
            task_order_no: 任务单号，由前端输入
            order_date: 制单日期，由前端输入
            delivery_date: 交货期，由前端输入
            output_dir: 输出目录，默认 output 文件夹
            pdf_name: PDF文件名
            customer_code: 客户号，由前端输入

        Returns:
            处理结果字典：
            双模板 {'GX': 'path1', 'GM': 'path2'}
            单模板 {'GX': 'path1'}
        """
        
        # 1. 区分双模板还是单模板
        template_number = len(template_2_info)

        # 单模板，只有广线。但要考虑是否有TL的情况；如果没TL，就是上一版本的处理，有TL，还需要处理 TL sheet
        if template_number == 1:
            gx_output_path = self.process_GX_template(raw_pdf_info, template_2_info, task_order_no, order_date, delivery_date, output_dir, pdf_name, customer_code, task_id)

            saved_path = {
                'GX': gx_output_path
            }

        # 双模板
        elif template_number == 2:
            """
            1. 区分从 pdf中 抽取处的产品 是广线还是广美
            2. 为了复用代码，将 raw_pdf_info 中 product_info 中仅保留对应的产品信息（如数量）；将最新的 new_pdf_info 传入
            3. 广美的直接用上一版处理代码；广线的直接调 template_number == 1 的代码
            """
            # 本次从 PO 中抽取的客户单号
            all_customer_codes_from_po = [p['cust_item_code'] for p in raw_pdf_info['product_info']]
            # 从 广线 模板中得到的所有客户单号
            all_customer_codes_from_gx_template = template_2_info['GX']['customer_codes']
            # 从 广美 模板中得到的所有客户单号
            all_customer_codes_from_gm_template = template_2_info['GM']['customer_codes']
            # 本次从 PO 中抽取的客户单号中有哪些属于 广美的产品
            gm_customer_codes_from_po = [c for c in all_customer_codes_from_po if c in all_customer_codes_from_gm_template]
            # 本次从 PO 中抽取的客户单号中有哪些属于 广线的产品
            gx_customer_codes_from_po = [c for c in all_customer_codes_from_po if c in all_customer_codes_from_gx_template]

            if len(gm_customer_codes_from_po) + len(gx_customer_codes_from_po) != len(raw_pdf_info['product_info']):
                raise ValueError("再划分PO单中的商品属于 广线 还是 广美 时出错！")

            # 仅保留 raw_pdf_info 中属于 广美的 客户单号，构造广美的 new_pdf_info，再传入 process_GM_template 方法
            new_pdf_info_from_gm = {
                "po_no": raw_pdf_info["po_no"],
                "product_info": [
                    item for item in raw_pdf_info["product_info"]
                    if item["cust_item_code"] in gm_customer_codes_from_po
                ]
            }
            gm_template_path = template_2_info['GM']['template_path']
            gm_wb = load_workbook(filename=gm_template_path, data_only=False)
            gm_output_path = self.process_GM_template(gm_wb, new_pdf_info_from_gm, order_date, delivery_date, task_order_no, customer_code, output_dir, template_2_info['GM']['customer_codes'], task_id)

            # 仅保留 raw_pdf_info 中属于 广线的 客户单号，构造广线的 new_pdf_info，在传入 process_GX_template 方法
            new_pdf_info_from_gx = {
                "po_no": raw_pdf_info["po_no"],
                "product_info": [
                    item for item in raw_pdf_info["product_info"]
                    if item["cust_item_code"] in gx_customer_codes_from_po
                ]
            }
            gx_output_path = self.process_GX_template(new_pdf_info_from_gx, template_2_info, task_order_no, order_date, delivery_date,
                                               output_dir, pdf_name, customer_code, task_id)

            saved_path = {
                'GX': gx_output_path,
                'GM': gm_output_path
            }

        else:
            raise Exception(f"模板数量有误：{template_number}，请检查！")

        return saved_path